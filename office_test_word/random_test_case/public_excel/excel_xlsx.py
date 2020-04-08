#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Author  : yag8009
# @FileName  : excel_xlsx
# @Time    : 2019/12/11

import openpyxl
from openpyxl.styles import PatternFill


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["case_id", "case_name", "url", "data", "assertion"])
    for i in range(1, index):
        # sheet.cell(row=i + 1, column=1,value=str(value[i]))
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


if __name__ == "__main__":
    book_name_xlsx = 'xlsx格式测试工作簿.xlsx'
    sheet_name_xlsx = 'xlsx格式测试表'
    value3 = [["姓名", "性别", "年龄", "城市", "职业"],
              ["111", "女", "66", "石家庄", "运维工程师"],
              ["222", "男", "55", "南京", "饭店老板"],
              ["333", "女", "27", "苏州", "保安"], ]
    write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Author  : yag8009
# @FileName  : yinke_shangpin_cmp
# @Time    : 2020/8/17
import re
import operator as op
import json_tools
import xlrd
from openpyxl import Workbook

from testcases.yinke_solr_API import yinke_solr_API


def yingke_request(url, data, key):
    result = yinke_solr_API(url=url, data=data, key=key).json()
    return result


def yingke_cmpvalue(src_data, dst_data):
    if isinstance(src_data, dict):
        """若为dict格式"""
        for key in dst_data:
            if key not in src_data:
                print("src不存在这个key")
        for key in src_data:
            if key in dst_data:
                thiskey = key
                """递归"""
                op.eq(src_data[key], dst_data[key])
            else:
                print("dst不存在这个key",key)
    elif isinstance(src_data, list):
        """若为list格式"""
        if len(src_data) != len(dst_data):
            print("list len: '{}' != '{}'".format(len(src_data), len(dst_data)))
        for src_list, dst_list in zip(sorted(src_data), sorted(dst_data)):
            """递归"""
            op.eq(src_list, dst_list)
    else:
        if str(src_data) != str(dst_data):
            print(src_data)
def sort_json(json_data):
    workflows = json_data["result"]
    return sorted(workflows, key=lambda d: d["id"])


def run():
    key = "KBmTWW0nvtC298rJ"
    solr_url = "http://116.196.102.10/svc-xls/a/basics/product/product/solrList?data="
    old_url = "http://116.196.102.10/svc-xls/a/basics/product/product/solrList?data="
    url_list = ["http://116.196.102.10/svc-xls/a/basics/product/product/solrList?data=",
                "http://116.196.102.10/svc-xls/a/basics/product/product/list?data="]

    raw = 2
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="请求数据")
    ws.cell(row=1, column=2, value="solr返回数据")
    ws.cell(row=1, column=3, value="老接口返回数据")
    ws.cell(row=1, column=4, value="新旧接口比较数据")
    datas = xlrd.open_workbook("F:\\solrnamedata.xlsx")
    # # 通过索引顺序获取Excel表
    table = datas.sheets()[0]
    for args in range(1, table.nrows):
        txt = table.row_values(args)
        data = {
            "method": "/basics/product/product/list",
            "companycode": "974277",
            "ktypeid": "51137",
            # "categoryId": txt[0],
            "searchKey": re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])", "", str(txt[0][:3])),
            "pageNo": "1",
            "pageSize": "999"
        }
        for ur in url_list:
            res = yingke_request(url=ur, data=data, key=key)
            print("返回数据: ", res)
            print("=" * 200)

            if ur in solr_url:
                solri = dict(res)
                news = sort_json(solri)
            else:
                oldi = dict(res)
                olds = sort_json(oldi)

        try:
            bukey = json_tools.diff(news, olds)
            ws.cell(row=raw, column=1, value=str(data))
            ws.cell(row=raw, column=2, value=str(solri))
            ws.cell(row=raw, column=3, value=str(oldi))
            ws.cell(row=raw, column=4, value=str(bukey))
            raw += 1
            wb.save('f:\example.xlsx')
            print("存储 bukey ：", bukey)
            print("=" * 200)
        except Exception as e:
            print(e)



if __name__ == '__main__':
    run()
